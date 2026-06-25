"""
Exports the Abeer HR sample data (shared with the PBIP build) into a typed
Excel workbook (AbeerHR_Data.xlsx) with one ListObject table per entity.
build_xlsm.ps1 then injects the VBA engine and saves the final .xlsm.
"""
import os, sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import build_pbip as src  # noqa: E402  (re-runs the deterministic data generation)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "AbeerHR_Data.xlsx")

def d(v):  # ISO string -> date (or None)
    return date.fromisoformat(v) if v else None

def num(v, f=int):
    return f(v) if v not in (None, "") else None

employees = [[e[0], e[1], e[2], e[3], e[4], e[5], e[6], e[7], d(e[8]), d(e[9]),
              e[10], e[11], num(e[12]), d(e[13]), e[14], e[15], e[16],
              num(e[17], float), e[18],
              "Active" if e[9] is None else "Terminated"] for e in src.employees]
recruitment = [[r[0], r[1], r[2], r[3], d(r[4]), d(r[5]), r[6], num(r[7]),
                num(r[8]), num(r[9]), num(r[10]), r[11]] for r in src.recruitment]
training = [[t[0], t[1], t[2], t[3], d(t[4]), num(t[5]), num(t[6]), t[7]] for t in src.training]
leave = [[l[0], l[1], l[2], d(l[3]), num(l[4])] for l in src.leave]

SHEETS = [
    ("Data_Employees", "tblEmployees",
     ["EmployeeID", "FullName", "Gender", "Nationality", "NationalityGroup", "Department",
      "JobTitle", "Facility", "HireDate", "TerminationDate", "TerminationReason",
      "TerminationType", "MonthlySalarySAR", "BirthDate", "AgeBand", "TenureBand",
      "Grade", "PerformanceRating", "EmploymentType", "Status"], employees),
    ("Data_Recruitment", "tblRecruitment",
     ["RequisitionID", "Position", "Department", "Facility", "PostingDate", "FilledDate",
      "Status", "DaysToFill", "Applications", "OffersMade", "OffersAccepted", "Source"], recruitment),
    ("Data_Training", "tblTraining",
     ["TrainingID", "EmployeeID", "Course", "Category", "TrainingDate", "Hours",
      "CostSAR", "CompletionStatus"], training),
    ("Data_Leave", "tblLeave",
     ["LeaveID", "EmployeeID", "LeaveType", "StartDate", "Days"], leave),
    ("Dim_Department", "tblDepartments", ["Department", "Division"], src.dim_department),
    ("Dim_Facility", "tblFacilities", ["Facility", "City", "Region", "FacilityType"], src.dim_facility),
    ("Config", "tblConfig", ["Key", "Value"],
     [["SaudizationTarget", 0.40], ["AttritionTargetMax", 0.15], ["TimeToFillTargetDays", 60],
      ["TrainingCompletionTarget", 0.90], ["AbsenteeismMax", 0.03], ["OpenReqAgeAlertDays", 60],
      ["SickDaysAlertThreshold", 10], ["LowRatingThreshold", 2.5]]),
]

wb = Workbook()
wb.remove(wb.active)
for sheet_name, table_name, headers, rows in SHEETS:
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for r in rows:
        ws.append(r)
    ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(28, len(h) + 4))
    # date display format
    for i, h in enumerate(headers, 1):
        if h.endswith("Date"):
            for c in ws[get_column_letter(i)][1:]:
                c.number_format = "DD-MMM-YYYY"

wb.save(OUT)
print(f"OK  {OUT}")
for s, t, h, r in SHEETS:
    print(f"    {s}: {len(r)} rows x {len(h)} cols ({t})")
